import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import openpyxl

def iter_pd(df):
    for val in df.columns:
        yield val
    for row in df.to_numpy():
        for val in row:
            if pd.isna(val):
                yield ""
            else:
                yield val

def pandas_to_sheets(pandas_df, sheet, clear = True):
    # Updates all values in a workbook to match a pandas dataframe
    if clear: sheet.clear()
    (row, col) = pandas_df.shape
    cells = sheet.range("A1:{}".format(gspread.utils.rowcol_to_a1(row + 1, col)))
    for cell, val in zip(cells, iter_pd(pandas_df)):
        cell.value = val
    sheet.update_cells(cells)

def update():
    # Connect to Google
    # Scope: Enable access to specific links
    scope = ['https://www.googleapis.com/auth/spreadsheets',
            "https://www.googleapis.com/auth/drive"]

    credentials = ServiceAccountCredentials.from_json_keyfile_name("gs_credentials.json", scope)
    client = gspread.authorize(credentials)

    # Create a blank spreadsheet (Note: We're using a service account, so this spreadsheet is visible only to this account)
    # sheet = client.open("K_K - Training")

    # To access newly created spreadsheet from Google Sheets with your own Google account you must share it with your email
    # Sharing a Spreadsheet
    # sheet.share('kareemkhaled143@gmail.com', perm_type='user', role='writer')

    # Open the spreadsheet
    sheet = client.open("K_K - Training")
    tasks = sheet.worksheet("Tasks")
    standing = sheet.worksheet("Standing")
    # read csv with pandas
    df = pd.ExcelFile('K_K - Training.xlsx')
    # export df to a sheet

    # try:print(df.values.tolist())
    # except: pass

    # try:print([df.columns.values.tolist()])
    # except: pass

    wb = openpyxl.load_workbook('K_K - Training.xlsx')
    ws = wb['Standing']
    # print(ws.cell(row=1, column=2).hyperlink.target)
    row = 0
    col = 1

    # for cell in ws.iter_rows():
    #     if row:
    #         standing.update_cell(row, 1, f'=HYPERLINK("{cell[0].hyperlink.target}", "{cell[0].value}")')
    #     row += 1
    pandas_to_sheets(pd.read_excel(df, 'Standing'), standing)

    for cell in ws.iter_cols():
        try:
            standing.update_cell(1, col, f'=HYPERLINK("{cell[0].hyperlink.target}", "{cell[0].value}")')
        except: pass
        col += 1


    # standing.update([pd.read_excel(df, 'Standing').columns.values.tolist()] + pd.read_excel(df, 'Standing').values.tolist())

    # print(tasks)
    # print(pd.read_excel(df, 'Standing'))
    # pandas_to_sheets(pd.read_excel(df, 'Tasks'), tasks)

if __name__ == '__main__':
    update()